# -*- coding: utf-8 -*-
"""
art-requirement skill — Excel 写入脚本

用法：
    python write_art_excel.py <json_path> <output_xlsx_path>

输入 JSON 结构：
{
  "feature_name": "功能名称（用于文件命名参考）",
  "animation": [ { 列名: 值, ... }, ... ],
  "vfx": [ { 列名: 值, ... }, ... ],
  "character_concept": [ { 列名: 值, ... }, ... ],
  "scene_concept": [ { 列名: 值, ... }, ... ]
}

输出：一个 .xlsx 文件，每类需求一个标签页。空数组的类别不创建标签页。
"""

import sys
import io
import json
import os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ─── Column definitions per sheet (order matters) ───

ANIMATION_COLUMNS = [
    "需求指派", "版本", "优先级", "动画类型", "动画中文名称",
    "制作进度", "动画资源命名", "备注", "动作参考",
    "动画时长", "是否循环", "融合状态", "连接动作",
    "制作进度", "制作美术", "提需策划", "上传路径",
]

VFX_COLUMNS = [
    "需求指派", "需求名", "优先级", "版本", "状态", "需求描述",
    "参考图片/视频", "参考图片/视频", "参考链接",
    "帕鲁资产位置", "最大视距", "是否循环", "持续时长",
    "特效资产命名", "资产上传路径", "策划", "备注", "父记录",
]

CONCEPT_ART_COLUMNS = [
    "需求指派", "版本", "状态", "需求名", "分类", "优先级",
    "需求描述", "参考图片", "参考图片", "需求文档",
    "策划", "制作人", "开始时间", "结束时间", "单号", "NAS地址", "父记录",
]

SHEET_CONFIG = {
    "animation":         ("动画需求", ANIMATION_COLUMNS),
    "vfx":               ("特效需求", VFX_COLUMNS),
    "character_concept": ("角色原画", CONCEPT_ART_COLUMNS),
    "scene_concept":     ("场景原画", CONCEPT_ART_COLUMNS),
}

# ─── Styling ───

HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def estimate_col_width(col_name, values):
    """Estimate column width based on header and content."""
    max_len = len(col_name) * 2
    for v in values[:20]:
        if v:
            max_len = max(max_len, min(len(str(v)) * 1.2, 60))
    return max(max_len, 8)


def write_sheet(wb, sheet_name, columns, rows):
    """Write one sheet with headers and data rows."""
    ws = wb.create_sheet(title=sheet_name)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    col_values = {c: [] for c in columns}
    for row_data in rows:
        for c in columns:
            col_values[c].append(row_data.get(c, ""))
    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (
            estimate_col_width(col_name, col_values[col_name])
        )

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    return len(rows)


def main():
    if len(sys.argv) < 3:
        print("用法: python write_art_excel.py <json_path> <output_xlsx_path>")
        sys.exit(1)

    json_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_rows = 0
    sheets_created = []

    for key, (sheet_name, columns) in SHEET_CONFIG.items():
        rows = data.get(key, [])
        if not rows:
            continue
        count = write_sheet(wb, sheet_name, columns, rows)
        total_rows += count
        sheets_created.append((sheet_name, count))

    if not sheets_created:
        print("WARNING: No data in any category, creating empty workbook.")
        wb.create_sheet(title="空")

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)

    print("Excel saved: %s" % os.path.abspath(output_path))
    for name, count in sheets_created:
        print("  Sheet [%s]: %d rows" % (name, count))
    print("Total: %d rows across %d sheets" % (total_rows, len(sheets_created)))


if __name__ == "__main__":
    main()
